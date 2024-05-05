from pyshadow.main import Shadow
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np
import shutil
import datetime
import math
from UliPlot.XLSX import auto_adjust_xlsx_column_width
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

now = str(datetime.datetime.now())[:19]
now = now.replace(":", "_")


def replace_nan_with_zero(value):
    return value if not math.isnan(value) else 0


# page to access as a string
url = 'https://www.knhb.nl/match-center#/competitions/N7/results'
driver = webdriver.Chrome()
driver.get(url)
driver.implicitly_wait(5)

cookies_popup = driver.find_element(By.XPATH, '//*[@id="bcSubmitConsentToAll"]')
if cookies_popup:
    driver.find_element(By.XPATH, '//*[@id="bcSubmitConsentToAll"]').click()
    driver.implicitly_wait(10)

shadow = Shadow(driver)
z = shadow.chrome_driver.get('https://www.knhb.nl/match-center#/competitions/N7/results')
element = shadow.find_element("match-center")
shadow.set_implicit_wait(10)
text = element.text
text = text.splitlines()
driver.close()

# months in Dutch
months = ["januari", 'februari', 'maart', 'april', 'mei', 'juni', 'juli', 'augustus', 'september', 'oktober',
          'november', 'december']
# removing dates
for i in text:
    for k in months:
        if k in i:
            text.remove(i)

text_odd = text[1::2]
text_even = text[0::2]

team_away = []
team_home = []
pool = []
score = []
split_scores = []
winner = []
goal_difference = []


def has_numbers(x):
    return any(char.isdigit() for char in x)


for s in text_even:
    if "H1" in s:
        team_home.append(s)
    elif len(s) == 1:
        pool.append(s)

for s in text_odd:
    if "H1" in s:
        team_away.append(s)
    elif has_numbers(s) and "-" in s and len(s) < 6:
        score.append(s)

# split the scores into home score and away score
for s in score:
    split_score = s.replace('-', ' ').split()
    split_scores.append(split_score)

split_scores = np.array(split_scores).flatten()
home_score = split_scores[0::2]
home_score = pd.to_numeric(home_score)
away_score = split_scores[1::2]
away_score = pd.to_numeric(away_score)
goal_difference = pd.to_numeric(goal_difference)

new_results = pd.DataFrame(data=[team_home, home_score, away_score, team_away, goal_difference, winner, pool]).T
new_results = new_results.rename(columns={0: 'Home Team', 1: 'Home Score', 2: 'Away Score', 3: 'Away Team',
                                          4: 'Goal Difference', 5: 'Winner', 6: 'Pool'})

new_results['Goal Difference'] = new_results['Home Score'] - new_results['Away Score']

new_results.loc[new_results['Goal Difference'] < 0, 'Winner'] = 'Away'
new_results.loc[new_results['Goal Difference'] == 0, 'Winner'] = 'Draw'
new_results.loc[new_results['Goal Difference'] > 0, 'Winner'] = 'Home'

# location of Excel file with results
file_path_little = r"C:\Users\Harry\OneDrive\Hockey\Results and Analysis\H1\H1_1K_results_2324.xlsx"
file_path_big = r"C:\Users\Harry Higgins\OneDrive\Hockey\Results and Analysis\H1\H1_1K_results_2324.xlsx"

# read the current scores off the Excel file
try:
    old_results = pd.read_excel(file_path_little, sheet_name='All Results')
    file_path = file_path_little
    dst_dir = r"C:\Users\Harry\OneDrive\Hockey\Results and Analysis\H1\Previous\H1_1K_results_2324_"+str(now)+".xlsx"
except IOError:
    old_results = pd.read_excel(file_path_big, sheet_name='All Results')
    file_path = file_path_big
    dst_dir = r"C:\Users\Harry Higgins\OneDrive\Hockey\Results and Analysis\
    H1\Previous\H1_1K_results_2324_"+str(now)+".xlsx"

# copy the previous Excel file to another folder for backup
shutil.copy(file_path, dst_dir)

# create a data frame of the old and new results
all_results = pd.concat([new_results, old_results])
# remove duplicate results
all_results = all_results.drop_duplicates()

with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    all_results.to_excel(writer, sheet_name='All Results', index=False)
    auto_adjust_xlsx_column_width(all_results, writer, sheet_name='All Results')

# creating the pools dataframes
pool_A = pd.DataFrame()
pool_B = pd.DataFrame()
pool_C = pd.DataFrame()
pool_D = pd.DataFrame()
home_results = pd.DataFrame()
away_results = pd.DataFrame()
home_vs_away_results = pd.DataFrame()

# for each team, run through all operations
for team in all_results['Home Team'].unique():

    # create a DataFrame of all team results Home and Away, side by side
    team_df = pd.concat([all_results[all_results['Home Team'] == team].reset_index(drop=True),
                         all_results[all_results['Away Team'] == team].reset_index(drop=True)], axis=1)

    # renaming all the columns of the data frame
    team_df = team_df.set_axis(['Home Team', 'Home Score', 'Away Score', 'Away Team', 'Goal Difference',
                                'Winner', 'Pool', 'home team', 'home score', 'away score', 'away team',
                                'goal difference', 'winner', 'pool'], axis=1)

    # calculate team goals scored, if NaN replace with 0
    total_goals_for = team_df['Home Score'].sum(skipna=True) + team_df['away score'].sum(skipna=True)
    replace_nan_with_zero(total_goals_for)

    # calculate team goals conceded, if NaN replace with 0
    total_goals_against = team_df['Away Score'].sum(skipna=True) + team_df['home score'].sum(skipna=True)
    replace_nan_with_zero(total_goals_against)

    # calculate team goal difference, if NaN replace with 0
    total_goal_difference = team_df['Goal Difference'].sum(skipna=True) - team_df['goal difference'].sum(skipna=True)
    replace_nan_with_zero(total_goal_difference)

    # calculate games played by team, if NaN replace with 0
    total_games_played = ((team_df[team_df['Home Team'] == team].shape[0]) +
                          (team_df[team_df['away team'] == team].shape[0]))
    replace_nan_with_zero(total_games_played)

    # calculate team wins, if NaN replace with 0
    wins = (team_df[team_df['Winner'] == 'Home'].shape[0]) + (team_df[team_df['winner'] == 'Away'].shape[0])
    replace_nan_with_zero(wins)

    # calculate team draws, if NaN replace with 0
    draws = (team_df[team_df['Winner'] == 'Draw'].shape[0]) + (team_df[team_df['winner'] == 'Draw'].shape[0])
    replace_nan_with_zero(draws)

    # calculate team losses, if NaN replace with 0
    losses = (team_df[team_df['Winner'] == 'Away'].shape[0]) + (team_df[team_df['winner'] == 'Home'].shape[0])
    replace_nan_with_zero(losses)

    # calculate team points, if NaN replace with 0
    points = (wins*3)+(draws*1)
    replace_nan_with_zero(points)

    # create the DataFrame of team results
    team_result = pd.DataFrame(data=[team, points, total_games_played, wins, draws, losses, total_goals_for,
                                     total_goals_against, total_goal_difference]).T

    # write the team results to their relevant sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        team_df.to_excel(writer, sheet_name=team, index=False)
        auto_adjust_xlsx_column_width(team_df, writer, sheet_name=team)

    # find which pool the specific team is in,
    # try the home pool column,
    # if not, then use the away pool column
    try:
        pool = team_df.at[0, 'Pool']
    except KeyError:
        pool = team_df.at[0, 'pool']

    # check which pool the team is in, concatenate that team's results to the overall pool results
    if pool == 'A':
        pool_A = pd.concat([pool_A, team_result])
        pool_result = pool_A
    elif pool == 'B':
        pool_B = pd.concat([pool_B, team_result])
        pool_result = pool_B
    elif pool == 'C':
        pool_C = pd.concat([pool_C, team_result])
        pool_result = pool_C
    else:
        pool_D = pd.concat([pool_D, team_result])
        pool_result = pool_D

    # rename the columns of the pool results
    pool_result = pool_result.rename(columns={0: 'Team', 1: 'Points', 2: 'Games Played', 3: 'Wins', 4: 'Draws',
                                              5: 'Losses', 6: 'Goals For', 7: 'Goals Against', 8: 'Goal Difference'})

    # write results for each pool to the correct sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        pool_result.to_excel(writer, sheet_name=pool, index=False)
        auto_adjust_xlsx_column_width(pool_result, writer, sheet_name=pool)

    # calculate home goals for, replace with 0 if NaN
    home_goals_for = team_df['Home Score'].sum(skipna=True)
    replace_nan_with_zero(home_goals_for)

    # calculate home goals against, replace with 0 if NaN
    home_goals_against = team_df['Away Score'].sum(skipna=True)
    replace_nan_with_zero(home_goals_against)

    # calculate home goal difference, replace with 0 if NaN
    home_goal_difference = team_df['Goal Difference'].sum(skipna=True)
    replace_nan_with_zero(home_goal_difference)

    # calculate games played at home, replace with 0 if NaN
    home_games_played = (team_df[team_df['Home Team'] == team].shape[0])
    replace_nan_with_zero(home_games_played)

    # calculate games won at home, replace with 0 if NaN
    home_wins = (team_df[team_df['Winner'] == 'Home'].shape[0])
    replace_nan_with_zero(home_wins)

    # calculate games drawn at home, replace with 0 if NaN
    home_draws = (team_df[team_df['Winner'] == 'Draw'].shape[0])
    replace_nan_with_zero(home_draws)

    # calculate losses at home, replace with 0 if NaN
    home_losses = (team_df[team_df['Winner'] == 'Away'].shape[0])
    replace_nan_with_zero(home_losses)

    # calculate points gained at home, replace with 0 if NaN
    home_points = (home_wins * 3) + (home_draws * 1)
    replace_nan_with_zero(home_points)

    # create DataFrame of home team results for this specific team
    home_team_result = pd.DataFrame(data=[team, home_points, home_games_played, home_wins, home_draws, home_losses,
                                          home_goals_for, home_goals_against, home_goal_difference]).T

    # rename columns of DataFrame
    home_team_result = home_team_result.rename(columns={0: 'Team', 1: 'Points', 2: 'Games Played', 3: 'Wins',
                                                        4: 'Draws', 5: 'Losses', 6: 'Goals For', 7: 'Goals Against',
                                                        8: 'Goal Difference'})

    # concatenate specific team results to DataFrame of all teams
    home_results = pd.concat([home_results, home_team_result], ignore_index=True)

    # write home results to correct Excel sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        home_results.to_excel(writer, sheet_name='Home Results', index=False)
        auto_adjust_xlsx_column_width(home_results, writer, sheet_name='Home Results')

    # calculate away goals for, replace with 0 if NaN
    away_goals_for = team_df['away score'].sum(skipna=True)
    replace_nan_with_zero(away_goals_for)

    # calculate away goals against, replace with 0 if NaN
    away_goals_against = team_df['home score'].sum(skipna=True)
    replace_nan_with_zero(away_goals_against)

    # calculate away goal difference, replace with 0 if NaN
    away_goal_difference = 0 - team_df['goal difference'].sum(skipna=True)
    replace_nan_with_zero(away_goal_difference)

    # calculate away games played, replace with 0 if NaN
    away_games_played = (team_df[team_df['away team'] == team].shape[0])
    replace_nan_with_zero(away_games_played)

    # calculate away wins, replace with 0 if NaN
    away_wins = (team_df[team_df['winner'] == 'Away'].shape[0])
    replace_nan_with_zero(away_wins)

    # calculate away draws, replace with 0 if NaN
    away_draws = (team_df[team_df['winner'] == 'Draw'].shape[0])
    replace_nan_with_zero(away_draws)

    # calculate away losses, replace with 0 if NaN
    away_losses = (team_df[team_df['winner'] == 'Home'].shape[0])
    replace_nan_with_zero(away_losses)

    # calculate away points, replace with 0 if NaN
    away_points = (away_wins * 3) + (away_draws * 1)
    replace_nan_with_zero(away_points)

    # create DataFrame of away result for specific team
    away_team_result = pd.DataFrame(data=[team, away_points, away_games_played, away_wins, away_draws, away_losses,
                                          away_goals_for, away_goals_against, away_goal_difference]).T

    # rename the columns in the specific team DataFrame
    away_team_result = away_team_result.rename(columns={0: 'Team', 1: 'Points', 2: 'Games Played', 3: 'Wins',
                                                        4: 'Draws', 5: 'Losses', 6: 'Goals For', 7: 'Goals Against',
                                                        8: 'Goal Difference'})
    # concatenate this specific team with all other teams
    away_results = pd.concat([away_results, away_team_result], ignore_index=True)

    # write the away results to teh specific Excel sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        away_results.to_excel(writer, sheet_name='Away Results', index=False)
        auto_adjust_xlsx_column_width(away_results, writer, sheet_name='Away Results')

    # calculate % of points gained at home, if variables are 0 then set equal to 0
    if home_points and points != 0:
        home_points_pcnt = home_points / points
    else:
        home_points_pcnt = 0

    # calculate % of points gained away, if variables are 0 then set equal to 0
    if away_points and points != 0:
        away_points_pcnt = away_points / points
    else:
        away_points_pcnt = 0

    # calculate % of goals scored at home, if variables are 0 then set equal to 0
    if home_goals_for and total_goals_for != 0:
        home_goals_for_pcnt = home_goals_for / total_goals_for
    else:
        home_goals_for_pcnt = 0

    # calculate % of goals scored away, if variables are 0 then set equal to 0
    if away_goals_for and total_goals_for != 0:
        away_goals_for_pcnt = away_goals_for / total_goals_for
    else:
        away_goals_for_pcnt = 0

    # calculate % of goals conceded at home, if variables are 0 then set equal to 0
    if home_goals_against and total_goals_against != 0:
        home_goals_against_pcnt = home_goals_against / total_goals_against
    else:
        home_goals_against_pcnt = 0

    # calculate % of goals conceded away, if variables are 0 then set equal to 0
    if away_goals_against and total_goals_against != 0:
        away_goals_against_pcnt = away_goals_against / total_goals_against
    else:
        away_goals_against_pcnt = 0

    # create DataFrame of Home vs Away results for specific team
    home_vs_away_team_result = pd.DataFrame(data=[team, total_games_played, home_points_pcnt, away_points_pcnt,
                                                  home_goals_for_pcnt, away_goals_for_pcnt, home_goals_against_pcnt,
                                                  away_goals_against_pcnt]).T

    # rename columns of DataFrame
    home_vs_away_team_result = home_vs_away_team_result.rename(columns={0: 'Team', 1: 'Games Played',
                                                                        2: '% Points Home',
                                                                        3: '% Points Away', 4: '% Goals Scored Home',
                                                                        5: '% Goals Scored Away',
                                                                        6: '% Goals Conceded Home',
                                                                        7: '% Goals Conceded Away'})

    # concatenate specific team result to DataFrame of all teams
    home_vs_away_results = pd.concat([home_vs_away_results, home_vs_away_team_result], ignore_index=True)

    percentile_rule = ColorScaleRule(
        start_type='percentile',
        start_value=10,
        start_color='ffaaaa',  # red-ish
        mid_type='num',
        mid_value=0,
        mid_color='aaffaa',  # green-ish
        end_type='percentile',
        end_value=90,
        end_color='ffaaaa')  # red-ish

    # custom named style for the index
    index_style = NamedStyle(name="Index Style", font=Font(color='000000', italic=False, bold=True),
                             alignment=Alignment(horizontal='left'))

    # write Home Vs Away DataFrame to specific Excel sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as xlsx:
        home_vs_away_results.to_excel(xlsx, sheet_name='Home Vs Away', index=False)
        # auto_adjust_xlsx_column_width(home_vs_away_results, writer, sheet_name='Home Vs Away')

        ws = xlsx.sheets['Home Vs Away']

        title_row = '1'
        value_cells = 'C1:{col}{row}'.format(col=get_column_letter(ws.max_column), row=ws.max_row)
        index_column = 'A'

        ws.column_dimensions[index_column].width = 21

        # color all value cells
        ws.conditional_formatting.add(value_cells, percentile_rule)

        for row in ws[value_cells]:
            for cell in row:
                cell.style = '0.00%'

        for cell in ws[index_column]:
            cell.style = index_style

        for cell in ws[title_row]:
            cell.style = 'Headline 1'

print("results uploaded")
