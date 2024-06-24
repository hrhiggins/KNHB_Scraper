import selenium.common.exceptions
from pyshadow.main import Shadow
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np
import shutil
import math
from UliPlot.XLSX import auto_adjust_xlsx_column_width
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import geopy.distance
import re


def general_scraper(url, file_path_little, file_path_big, dst_dir_little, dst_dir_big,
                    club_location_little, club_location_big, even):

    # function to set NaN values with 0
    def replace_nan_with_zero(value):
        return value if not math.isnan(value) else 0

    def multiple_replace(string, rep_dict):
        pattern = re.compile("|".join([re.escape(k) for k in sorted(rep_dict, key=len, reverse=True)]), flags=re.DOTALL)
        return pattern.sub(lambda x: rep_dict[x.group(0)], string)

    # function to check if an array string contains digits
    def has_numbers(x):
        return any(char.isdigit() for char in x)

    test = True
    sex = ['H1', 'D1']

    # repeat until it works
    while test:
        try:
            # page to access as a string
            driver = webdriver.Chrome()
            driver.get(url)
            # wait if page has not loaded
            driver.implicitly_wait(5)

            # deal with cookies popup
            cookies_popup = driver.find_element(By.XPATH, '//*[@id="bcSubmitConsentToAll"]')
            if cookies_popup:
                driver.find_element(By.XPATH, '//*[@id="bcSubmitConsentToAll"]').click()
                driver.implicitly_wait(10)

            # using the ShadowDriver, find the element with the scores
            shadow = Shadow(driver)
            element = shadow.find_element("match-center")
            # wait for the element to load
            shadow.set_implicit_wait(10)
            text = element.text
            # split each text string into a new line
            text = text.splitlines()
            # close the webdriver
            driver.close()
            print('Success scraping from URL')
            test = False

        except selenium.common.exceptions.ElementNotVisibleException:
            print('Error scraping from URL')
            test = True

    # read the current scores off the Excel file, check both places it could exist, locate folder to place backup
    try:
        old_results = pd.read_excel(file_path_little, sheet_name='All Results')
        file_path = file_path_little
        dst_dir = dst_dir_little
        club_locations = pd.read_excel(club_location_little, sheet_name='clubs')
    except IOError:
        old_results = pd.read_excel(file_path_big, sheet_name='All Results')
        file_path = file_path_big
        dst_dir = dst_dir_big
        club_locations = pd.read_excel(club_location_big, sheet_name='clubs')

    # months in Dutch
    months = ["januari", 'februari', 'maart', 'april', 'mei', 'juni', 'juli', 'augustus', 'september', 'oktober',
              'november', 'december']
    # removing dates for the data
    for i in text:
        for k in months:
            if k in i:
                text.remove(i)

    # split the data into even position and odd position
    text_odd = text[1::2]
    text_even = text[0::2]

    # creating all the used arrays
    team_away = []
    team_home = []
    pool = []
    score = []
    split_scores = []
    winner = []
    goal_difference = []
    distance = []

    if even == "True":
        # filter even text into team name and pool
        for s in text_even:
            if any(x in s for x in sex):
                team_home.append(s)
            elif len(s) == 1:
                pool.append(s)

        # filter odd text into away team and score
        for s in text_odd:
            if any(x in s for x in sex):
                team_away.append(s)
            elif has_numbers(s) and "-" in s and len(s) < 6:
                score.append(s)
            elif s == 'Afgelast':
                score.append('69 - 69')

    elif even == "False":
        # filter even text into team name and pool
        for s in text_odd:
            if any(x in s for x in sex):
                team_home.append(s)
            elif len(s) == 1:
                pool.append(s)

        # filter odd text into away team and score
        for s in text_even:
            if any(x in s for x in sex):
                team_away.append(s)
            elif has_numbers(s) and "-" in s and len(s) < 6:
                score.append(s)
            elif s == 'Afgelast':
                score.append('69 - 69')

    # split the scores into home score and away score
    for s in score:
        split_score = s.replace('-', ' ').split()
        split_scores.append(split_score)

    # filter the scores into a home score and an away score
    split_scores = np.array(split_scores).flatten()
    home_score = split_scores[0::2]
    home_score = pd.to_numeric(home_score)
    away_score = split_scores[1::2]
    away_score = pd.to_numeric(away_score)
    goal_difference = pd.to_numeric(goal_difference)

    # create a DataFrame comprised of the new downloaded data
    new_results = pd.DataFrame(data=[team_home, home_score, away_score, team_away, goal_difference, winner, pool,
                                     distance]).T
    new_results = new_results.rename(columns={0: 'Home Team', 1: 'Home Score', 2: 'Away Score', 3: 'Away Team',
                                              4: 'Goal Difference', 5: 'Winner', 6: 'Pool', 7: 'Distance'})

    # calculate the goal difference stat
    new_results['Goal Difference'] = new_results['Home Score'] - new_results['Away Score']

    for s in new_results['Home Team']:
        for i in sex:
            s = s.replace(i, '')
        try:
            club_locations['Club'] == s
        except KeyError:
            print(s, 'is not listed in club locations document')
            exit()

    for i in range(len(new_results)):
        h_team = new_results['Home Team'].iloc[i]
        h_club = h_team.replace('H1', '').replace('D1', '')
        h_location = club_locations[club_locations['Club'] == h_club]['Location']
        a_team = new_results['Away Team'].iloc[i]
        a_club = a_team.replace('H1', '').replace('D1', '')
        a_location = club_locations[club_locations['Club'] == a_club]['Location']
        distance = round(geopy.distance.geodesic(h_location, a_location).km, 1)
        new_results['Distance'].iloc[i] = distance

    # find out who won the game, based on goal difference
    new_results.loc[new_results['Goal Difference'] < 0, 'Winner'] = 'Away'
    new_results.loc[new_results['Goal Difference'] == 0, 'Winner'] = 'Draw'
    new_results.loc[new_results['Goal Difference'] > 0, 'Winner'] = 'Home'

    # copy the previous Excel file to another folder for backup
    shutil.copy(file_path, dst_dir)

    # create a data frame of the old and new results
    all_results = pd.concat([new_results, old_results])
    # remove duplicate results
    all_results = all_results.drop_duplicates()

    # drop cancelled games where the scores of the games have been changed to 69 - 69
    all_results = all_results.drop(all_results[(all_results['Home Score'] == 69) & (all_results['Away Score'] == 69)].
                                   index)

    # write to All Results Excel Sheet
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        all_results.to_excel(writer, sheet_name='All Results', index=False)
        auto_adjust_xlsx_column_width(all_results, writer, sheet_name='All Results')

    # creating all the dataframes used for each team
    pool_A = pd.DataFrame()
    pool_B = pd.DataFrame()
    pool_C = pd.DataFrame()
    pool_D = pd.DataFrame()
    home_results = pd.DataFrame()
    away_results = pd.DataFrame()
    home_vs_away_results = pd.DataFrame()
    relative_home_away = pd.DataFrame()

    # for each team, run through all operations
    for team in all_results['Home Team'].unique():

        # create a DataFrame of all team results Home and Away, side by side
        team_df = pd.concat([all_results[all_results['Home Team'] == team].reset_index(drop=True),
                             all_results[all_results['Away Team'] == team].reset_index(drop=True)], axis=1)

        # renaming all the columns of the data frame
        team_df = team_df.set_axis(['Home Team', 'Home Score', 'Away Score', 'Away Team', 'Goal Difference',
                                    'Winner', 'Pool', 'Distance', 'home team', 'home score', 'away score', 'away team',
                                    'goal difference', 'winner', 'pool', 'distance'], axis=1)

        # calculate team goals scored, if NaN replace with 0
        total_goals_for = team_df['Home Score'].sum(skipna=True) + team_df['away score'].sum(skipna=True)
        replace_nan_with_zero(total_goals_for)

        # calculate team goals conceded, if NaN replace with 0
        total_goals_against = team_df['Away Score'].sum(skipna=True) + team_df['home score'].sum(skipna=True)
        replace_nan_with_zero(total_goals_against)

        # calculate team goal difference, if NaN replace with 0
        total_goal_difference = team_df['Goal Difference'].sum(skipna=True) - team_df['goal difference'].sum(
            skipna=True)
        replace_nan_with_zero(total_goal_difference)

        # calculate games played by team, if NaN replace with 0
        total_games_played = ((team_df[team_df['Home Team'] == team].shape[0]) +
                              (team_df[team_df['away team'] == team].shape[0]))
        replace_nan_with_zero(total_games_played)

        # calculate team wins, if NaN replace with 0
        wins = (team_df[team_df['Winner'] == 'Home'].shape[0]) + (team_df[team_df['winner'] == 'Away'].shape[0])
        replace_nan_with_zero(wins)

        # calculate team draws, if NaN replace with 0
        draws = (team_df[team_df['Winner'] == 'Draw'].shape[0]) + (team_df[team_df['winner'] == 'Draw'].shape[0])
        replace_nan_with_zero(draws)

        # calculate team losses, if NaN replace with 0
        losses = (team_df[team_df['Winner'] == 'Away'].shape[0]) + (team_df[team_df['winner'] == 'Home'].shape[0])
        replace_nan_with_zero(losses)

        # calculate team points, if NaN replace with 0
        points = (wins * 3) + (draws * 1)
        replace_nan_with_zero(points)

        # create the DataFrame of team results
        team_result = pd.DataFrame(data=[team, points, total_games_played, wins, draws, losses, total_goals_for,
                                         total_goals_against, total_goal_difference]).T

        # write the team results to their relevant sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            team_df.to_excel(writer, sheet_name=team, index=False)
            auto_adjust_xlsx_column_width(team_df, writer, sheet_name=team)

        # find which pool the specific team is in,
        # try the home pool column,
        # if not, then use the away pool column
        try:
            pool = team_df.at[0, 'Pool']
        except KeyError:
            pool = team_df.at[0, 'pool']

        # check which pool the team is in, concatenate that team's results to the overall pool results
        if pool == 'A':
            pool_A = pd.concat([pool_A, team_result])
            pool_result = pool_A
        elif pool == 'B':
            pool_B = pd.concat([pool_B, team_result])
            pool_result = pool_B
        elif pool == 'C':
            pool_C = pd.concat([pool_C, team_result])
            pool_result = pool_C
        else:
            pool_D = pd.concat([pool_D, team_result])
            pool_result = pool_D

        # rename the columns of the pool results
        pool_result = pool_result.rename(columns={0: 'Team', 1: 'Points', 2: 'Games Played', 3: 'Wins', 4: 'Draws',
                                                  5: 'Losses', 6: 'Goals For', 7: 'Goals Against',
                                                  8: 'Goal Difference'})

        # write results for each pool to the correct sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pool_result.to_excel(writer, sheet_name=pool, index=False)
            auto_adjust_xlsx_column_width(pool_result, writer, sheet_name=pool)

        # calculate home goals for, replace with 0 if NaN
        home_goals_for = team_df['Home Score'].sum(skipna=True)
        replace_nan_with_zero(home_goals_for)

        # calculate home goals against, replace with 0 if NaN
        home_goals_against = team_df['Away Score'].sum(skipna=True)
        replace_nan_with_zero(home_goals_against)

        # calculate home goal difference, replace with 0 if NaN
        home_goal_difference = team_df['Goal Difference'].sum(skipna=True)
        replace_nan_with_zero(home_goal_difference)

        # calculate games played at home, replace with 0 if NaN
        home_games_played = (team_df[team_df['Home Team'] == team].shape[0])
        replace_nan_with_zero(home_games_played)

        # calculate games won at home, replace with 0 if NaN
        home_wins = (team_df[team_df['Winner'] == 'Home'].shape[0])
        replace_nan_with_zero(home_wins)

        # calculate games drawn at home, replace with 0 if NaN
        home_draws = (team_df[team_df['Winner'] == 'Draw'].shape[0])
        replace_nan_with_zero(home_draws)

        # calculate losses at home, replace with 0 if NaN
        home_losses = (team_df[team_df['Winner'] == 'Away'].shape[0])
        replace_nan_with_zero(home_losses)

        # calculate points gained at home, replace with 0 if NaN
        home_points = (home_wins * 3) + (home_draws * 1)
        replace_nan_with_zero(home_points)

        # create DataFrame of home team results for this specific team
        home_team_result = pd.DataFrame(data=[team, home_points, home_games_played, home_wins, home_draws, home_losses,
                                              home_goals_for, home_goals_against, home_goal_difference]).T

        # rename columns of DataFrame
        home_team_result = home_team_result.rename(columns={0: 'Team', 1: 'Points', 2: 'Games Played', 3: 'Wins',
                                                            4: 'Draws', 5: 'Losses', 6: 'Goals For', 7: 'Goals Against',
                                                            8: 'Goal Difference'})

        # concatenate specific team results to DataFrame of all teams
        home_results = pd.concat([home_results, home_team_result], ignore_index=True)

        # write home results to correct Excel sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            home_results.to_excel(writer, sheet_name='Home Results', index=False)
            auto_adjust_xlsx_column_width(home_results, writer, sheet_name='Home Results')

        # calculate away goals for, replace with 0 if NaN
        away_goals_for = team_df['away score'].sum(skipna=True)
        replace_nan_with_zero(away_goals_for)

        # calculate away goals against, replace with 0 if NaN
        away_goals_against = team_df['home score'].sum(skipna=True)
        replace_nan_with_zero(away_goals_against)

        # calculate away goal difference, replace with 0 if NaN
        away_goal_difference = 0 - team_df['goal difference'].sum(skipna=True)
        replace_nan_with_zero(away_goal_difference)

        # calculate away games played, replace with 0 if NaN
        away_games_played = (team_df[team_df['away team'] == team].shape[0])
        replace_nan_with_zero(away_games_played)

        # calculate away wins, replace with 0 if NaN
        away_wins = (team_df[team_df['winner'] == 'Away'].shape[0])
        replace_nan_with_zero(away_wins)

        # calculate away draws, replace with 0 if NaN
        away_draws = (team_df[team_df['winner'] == 'Draw'].shape[0])
        replace_nan_with_zero(away_draws)

        # calculate away losses, replace with 0 if NaN
        away_losses = (team_df[team_df['winner'] == 'Home'].shape[0])
        replace_nan_with_zero(away_losses)

        # calculate away points, replace with 0 if NaN
        away_points = (away_wins * 3) + (away_draws * 1)
        replace_nan_with_zero(away_points)

        # create DataFrame of away result for specific team
        away_team_result = pd.DataFrame(data=[team, away_points, away_games_played, away_wins, away_draws, away_losses,
                                              away_goals_for, away_goals_against, away_goal_difference]).T

        # rename the columns in the specific team DataFrame
        away_team_result = away_team_result.rename(columns={0: 'Team', 1: 'Points', 2: 'Games Played', 3: 'Wins',
                                                            4: 'Draws', 5: 'Losses', 6: 'Goals For', 7: 'Goals Against',
                                                            8: 'Goal Difference'})

        # concatenate this specific team with all other teams
        away_results = pd.concat([away_results, away_team_result], ignore_index=True)

        # write the away results to teh specific Excel sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            away_results.to_excel(writer, sheet_name='Away Results', index=False)
            auto_adjust_xlsx_column_width(away_results, writer, sheet_name='Away Results')

        # calculate % of points gained at home, if variables are 0 then set equal to 0
        if home_points and points != 0:
            home_points_pcnt = home_points / points
        else:
            home_points_pcnt = 0

        # calculate % of points gained away, if variables are 0 then set equal to 0
        if away_points and points != 0:
            away_points_pcnt = away_points / points
        else:
            away_points_pcnt = 0

        # calculate % of goals scored at home, if variables are 0 then set equal to 0
        if home_goals_for and total_goals_for != 0:
            home_goals_for_pcnt = home_goals_for / total_goals_for
        else:
            home_goals_for_pcnt = 0

        # calculate % of goals scored away, if variables are 0 then set equal to 0
        if away_goals_for and total_goals_for != 0:
            away_goals_for_pcnt = away_goals_for / total_goals_for
        else:
            away_goals_for_pcnt = 0

        # calculate % of goals conceded at home, if variables are 0 then set equal to 0
        if home_goals_against and total_goals_against != 0:
            home_goals_against_pcnt = home_goals_against / total_goals_against
        else:
            home_goals_against_pcnt = 0

        # calculate % of goals conceded away, if variables are 0 then set equal to 0
        if away_goals_against and total_goals_against != 0:
            away_goals_against_pcnt = away_goals_against / total_goals_against
        else:
            away_goals_against_pcnt = 0

        # create DataFrame of Home vs Away results for specific team
        home_vs_away_team_result = pd.DataFrame(data=[team, pool, total_games_played, points, home_points_pcnt,
                                                      away_points_pcnt, home_goals_for_pcnt, away_goals_for_pcnt,
                                                      home_goals_against_pcnt, away_goals_against_pcnt]).T

        # rename columns of DataFrame
        home_vs_away_team_result = home_vs_away_team_result.rename(columns={0: 'Team', 1: 'Pool', 2: 'Games Played',
                                                                            3: 'Points', 4: 'Points Home',
                                                                            5: 'Points Away', 6: 'Goals Scored Home',
                                                                            7: 'Goals Scored Away',
                                                                            8: 'Goals Conceded Home',
                                                                            9: 'Goals Conceded Away'})

        # concatenate specific team result to DataFrame of all teams
        home_vs_away_results = pd.concat([home_vs_away_results, home_vs_away_team_result], ignore_index=True)

        # define the percentile rule to colour the stats
        percentile_rule = ColorScaleRule(
            start_type='percent',
            start_value=35,
            start_color='ffaaaa',  # red-ish
            mid_type='percent',
            mid_value=50,
            mid_color='aaffaa',  # green-ish
            end_type='percent',
            end_value=65,
            end_color='ffaaaa')  # red-ish

        # establish the size of different cell borders
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="000000")

        # write Home Vs Away DataFrame to specific Excel sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            home_vs_away_results.to_excel(writer, sheet_name='Home Vs Away', index=False)
            auto_adjust_xlsx_column_width(home_vs_away_results, writer, sheet_name='Home Vs Away')

            ws = writer.sheets['Home Vs Away']

            # define what areas of the sheet I will alter
            title_row = '1'
            value_cells = 'E1:{col}{row}'.format(col=get_column_letter(ws.max_column), row=ws.max_row)
            index_column = 'A'

            # set the dimensions of the first (team) column
            ws.column_dimensions[index_column].width = 21

            # color all value cells
            ws.conditional_formatting.add(value_cells, percentile_rule)

            # format the bulk of the data
            for row in ws[value_cells]:
                for cell in row:
                    cell.number_format = '0.0%'
                    cell.border = Border(top=thin, left=double, right=double, bottom=thin)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # formate the title row
            for cell in ws[title_row]:
                cell.style = 'Headline 1'
                cell.border = Border(top=double, left=double, right=double, bottom=double)
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor="BDD7EE")

        # calculate ppg_home, check if variables are real
        if home_points and home_games_played != 0:
            ppg_home = home_points / home_games_played
        else:
            ppg_home = 0

        # calculate ppg_away, check if variables are real
        if away_points and away_games_played != 0:
            ppg_away = away_points / away_games_played
        else:
            ppg_away = 0

        # calculate ppg_difference, check if variables are real
        if ppg_home and ppg_away != 0:
            ppg_difference = ppg_home - ppg_away
        else:
            ppg_difference = 0

        # create dataframe of relative results
        relative_home_away_team = pd.DataFrame(data=[team, pool, home_games_played, away_games_played, points, ppg_home,
                                                     ppg_away, ppg_difference]).T

        # rename dataframe columns
        relative_home_away_team = relative_home_away_team.rename(columns={0: 'Team', 1: 'Pool', 2: 'GPh', 3: 'GPa',
                                                                          4: 'Points',
                                                                          5: 'PPG Home', 6: 'PPG Away',
                                                                          7: 'PPG Difference'})

        # add new relative results to already calculated ones
        relative_home_away = pd.concat([relative_home_away, relative_home_away_team], ignore_index=True)

        # define the color scale rule for the ppg difference
        difference_rule = ColorScaleRule(
            start_type='num',
            start_value=-1.2,
            start_color='ffaaaa',  # red-ish
            mid_type='num',
            mid_value=0,
            mid_color='aaffaa',  # green-ish
            end_type='num',
            end_value=1.2,
            end_color='ffaaaa')  # red-ish

        # write relative home away stats to correct sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            relative_home_away.to_excel(writer, sheet_name='Relative Home Away', index=False)

            # auto adjust column width
            auto_adjust_xlsx_column_width(relative_home_away, writer, sheet_name='Relative Home Away')

            ws = writer.sheets['Relative Home Away']

            # define which sets of cells I want to format
            title_row = '1'
            index_column = 'A'
            general_cells = 'B1:{col}{row}'.format(col=get_column_letter(ws.max_column), row=ws.max_row)
            ppg_difference_cells = 'H1:{col}{row}'.format(col=get_column_letter(ws.max_column), row=ws.max_row)

            # set with of Teams column
            ws.column_dimensions[index_column].width = 21

            # define formatting of the general cells
            for row in ws[general_cells]:
                for cell in row:
                    cell.number_format = '0.00'
                    cell.border = Border(top=thin, left=double, right=double, bottom=thin)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # format the ppg difference cells with the correct color scale rule
            ws.conditional_formatting.add(ppg_difference_cells, difference_rule)

            # format the title row
            for cell in ws[title_row]:
                cell.style = 'Headline 1'
                cell.border = Border(top=double, left=double, right=double, bottom=double)
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor="BDD7EE")

        ppk_difference_rule = ColorScaleRule(
            start_type='num',
            start_value=-1.2,
            start_color='ffaaaa',  # red-ish
            mid_type='num',
            mid_value=0,
            mid_color='aaffaa',  # green-ish
            end_type='num',
            end_value=1.2,
            end_color='ffaaaa')  # red-ish

        # calculate total distance travelled
        total_distance = team_df['distance'].sum(skipna=True)
        try:
            points_per_km = (away_points / total_distance)*100
        except ZeroDivisionError:
            points_per_km = 0

        team_ppk_results = pd.DataFrame(data=[team, points, total_games_played, away_points, total_distance, points_per_km]).T
        team_ppk_results = team_ppk_results.rename(columns={0: 'Team', 1: 'Points', 2: 'Games Played', 3: 'Away Points',
                                                                          4: 'Total Distance',
                                                                          5: 'Points per kM Travelled'})

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            team_ppk_results.to_excel(writer, sheet_name='Points Per kM', index=False)

            # auto adjust column width
            auto_adjust_xlsx_column_width(team_ppk_results, writer, sheet_name='Points Per kM')

            ws = writer.sheets['Points Per kM']

            # define which sets of cells I want to format
            title_row = '1'
            index_column = 'A'
            general_cells = 'B1:{col}{row}'.format(col=get_column_letter(ws.max_column), row=ws.max_row)
            ppk_difference_cells = 'F1:{col}{row}'.format(col=get_column_letter(ws.max_column), row=ws.max_row)

            # set with of Teams column
            ws.column_dimensions[index_column].width = 21

            # define formatting of the general cells
            for row in ws[general_cells]:
                for cell in row:
                    cell.number_format = '0.00'
                    cell.border = Border(top=thin, left=double, right=double, bottom=thin)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # format the ppg difference cells with the correct color scale rule
            ws.conditional_formatting.add(ppk_difference_cells, difference_rule)

            # format the title row
            for cell in ws[title_row]:
                cell.style = 'Headline 1'
                cell.border = Border(top=double, left=double, right=double, bottom=double)
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor="BDD7EE")





    # show that the results are uploaded in the console
    print("results uploaded")
